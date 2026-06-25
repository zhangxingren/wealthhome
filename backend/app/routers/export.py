"""数据导出 — Excel / CSV 分类导出"""

import csv
from io import BytesIO, StringIO
from typing import Optional, List
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.database import get_db
from app.auth import get_current_user

router = APIRouter(prefix="/api/export", tags=["导出"])

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="微软雅黑", size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _style_sheet(ws, headers: list, rows: list, col_widths: list = None):
    """统一设置表头样式和列宽"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    for r_idx, row in enumerate(rows, 2):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="center")

    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
    ws.auto_filter.ref = ws.dimensions


@router.get("/excel")
def export_excel(
    categories: Optional[str] = Query(None, description="逗号分隔分类: cash,deposit,fund,stock,bond,precious_metal,liability，不传则全部"),
    user=Depends(get_current_user),
):
    """导出资产/负债为 Excel"""
    uid = int(user["sub"])
    selected = set(categories.split(",")) if categories else None

    wb = Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)
    sheets_added = 0

    def _add_sheet(title, sql, headers, col_widths, row_mapper):
        nonlocal sheets_added
        with get_db() as db:
            rows = db.execute(sql, (uid,)).fetchall()
        if not rows:
            return
        ws = wb.create_sheet(title=title)
        _style_sheet(ws, headers, [row_mapper(r) for r in rows], col_widths)
        sheets_added += 1

    SHEET_CONFIGS = [
        ("cash",    "现金活期", "SELECT name,currency,amount,account_name,note,tags FROM asset_cash WHERE user_id=? ORDER BY id", ["名称","币种","金额","账户","备注","标签"], [16,8,14,14,20,14], lambda r: [r[i] for i in range(6)]),
        ("deposit", "定期存单", "SELECT name,bank,principal,rate,start_date,end_date,currency,note,tags FROM asset_deposit WHERE user_id=? ORDER BY id", ["名称","银行","本金","年利率(%)","起息日","到期日","币种","备注","标签"], [16,14,14,12,12,12,8,20,14], lambda r: [r[i] for i in range(9)]),
        ("fund",    "基金持仓", "SELECT code,name,shares,cost_nav,current_nav,fund_type,note,tags FROM asset_fund WHERE user_id=? ORDER BY id", ["基金代码","名称","份额","成本净值","当前净值","类型","备注","标签"], [12,16,12,12,12,10,20,14], lambda r: [r[i] for i in range(8)]),
        ("stock",   "股票持仓", "SELECT code,name,shares,cost_price,current_price,market,note,tags FROM asset_stock WHERE user_id=? ORDER BY id", ["股票代码","名称","持股数","成本价","当前价","市场","备注","标签"], [12,16,12,12,12,8,20,14], lambda r: [r[i] for i in range(8)]),
        ("bond",    "债权",    "SELECT name,issuer,face_value,rate,maturity_date,currency,note,tags FROM asset_bond WHERE user_id=? ORDER BY id", ["名称","发行方","面值","年利率(%)","到期日","币种","备注","标签"], [16,16,14,12,12,8,20,14], lambda r: [r[i] for i in range(8)]),
        ("precious_metal","贵金属","SELECT name,type,weight_grams,buy_price_per_gram,current_price_per_gram,buy_date,buy_total,notes FROM precious_metals WHERE user_id=? ORDER BY id", ["名称","类型","克重","买入价/克","现价/克","买入日","买入总价","备注"], [16,10,10,12,12,12,14,20], lambda r: [r[0], dict(zip(["gold","silver","platinum","palladium"],["黄金","白银","铂金","钯金"])).get(r[1],r[1]), *r[2:]]),
        ("liability","负债",   "SELECT name,principal,rate,term_months,repay_type,start_date,monthly_payment,remaining,note FROM liabilities WHERE user_id=? ORDER BY id", ["名称","本金","年利率(%)","期限(月)","还款方式","起始日","月供","剩余本金","备注"], [16,14,12,10,12,12,12,14,20], lambda r: [r[i] for i in range(9)]),
    ]

    for cat_key, title, sql, headers, widths, mapper in SHEET_CONFIGS:
        if not selected or cat_key in selected:
            _add_sheet(title, sql, headers, widths, mapper)

    if sheets_added == 0:
        ws = wb.create_sheet("无数据")
        ws.cell(row=1, column=1, value="暂无资产数据")

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = "wealthhome_export.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/csv")
def export_csv(
    categories: Optional[str] = Query(None, description="逗号分隔分类，不传则导出全部"),
    user=Depends(get_current_user),
):
    """导出资产/负债为 CSV"""
    uid = int(user["sub"])
    selected = set(categories.split(",")) if categories else None

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["分类", "名称/代码", "金额/份额", "单位", "备注"])

    CATEGORY_CONFIGS = [
        ("cash",    "现金活期", "SELECT name,amount,currency,note FROM asset_cash WHERE user_id=? ORDER BY id", lambda r: [r[0], str(r[1]), r[2], r[3]]),
        ("deposit", "定期存单", "SELECT name,principal,rate,currency,note FROM asset_deposit WHERE user_id=? ORDER BY id", lambda r: [r[0], f"{r[1]} (利率{r[2]}%)", r[3], r[4]]),
        ("fund",    "基金持仓", "SELECT code,name,shares,current_nav,note FROM asset_fund WHERE user_id=? ORDER BY id", lambda r: [f"{r[0]} {r[1]}", f"{r[2]} 份 × {r[3]} 净值", "CNY", r[4]]),
        ("stock",   "股票持仓", "SELECT code,name,shares,current_price,note FROM asset_stock WHERE user_id=? ORDER BY id", lambda r: [f"{r[0]} {r[1]}", f"{r[2]} 股 × {r[3]} 元", "CNY", r[4]]),
        ("bond",    "债权",     "SELECT name,issuer,face_value,rate,currency,note FROM asset_bond WHERE user_id=? ORDER BY id", lambda r: [f"{r[0]} ({r[1]})", f"{r[2]} (利率{r[3]}%)", r[4], r[5]]),
        ("precious_metal","贵金属","SELECT name,type,weight_grams,current_price_per_gram,notes FROM precious_metals WHERE user_id=? ORDER BY id", lambda r: [f"{r[0]} ({dict(zip(['gold','silver','platinum','palladium'],['黄金','白银','铂金','钯金'])).get(r[1],r[1])})", f"{r[2]} 克 × {r[3]} 元/克", "CNY", r[4]]),
        ("liability","负债",    "SELECT name,principal,rate,remaining,note FROM liabilities WHERE user_id=? ORDER BY id", lambda r: [r[0], f"{r[1]} (利率{r[2]}%, 剩余{r[3]})", "CNY", r[4]]),
    ]

    for cat_key, cat_label, sql, mapper in CATEGORY_CONFIGS:
        if not selected or cat_key in selected:
            with get_db() as db:
                rows = db.execute(sql, (uid,)).fetchall()
            for row in rows:
                mapped = mapper(list(row))
                writer.writerow([cat_label, mapped[0], mapped[1], mapped[2], mapped[3]])

    output.seek(0)
    filename = "wealthhome_export.csv"
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv; charset=utf-8-sig",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/json")
def export_json(categories: Optional[str] = Query(None), user=Depends(get_current_user)):
    """导出全部数据为 JSON（备份用）"""
    uid = int(user["sub"])
    selected = set(categories.split(",")) if categories else None
    data = {}
    with get_db() as db:
        tables = {
            "cash": "asset_cash",
            "deposit": "asset_deposit",
            "fund": "asset_fund",
            "stock": "asset_stock",
            "bond": "asset_bond",
            "precious_metal": "precious_metals",
            "liabilities": "liabilities",
        }
        for key, table in tables.items():
            if selected and key not in selected:
                continue
            rows = db.execute(f"SELECT * FROM {table} WHERE user_id=? ORDER BY id", (uid,)).fetchall()
            data[key] = [dict(r) for r in rows]
    return data
